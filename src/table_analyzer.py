# -*- coding: utf-8 -*-
"""表格分析模块：读取 Excel 并转换为纯文本格式"""
import openpyxl
from data_link import get_table_path


def _get_cell_value(cell):
    """安全获取单元格值，跳过 MergedCell"""
    try:
        return cell.value
    except AttributeError:
        return None


def _read_sheet_data(ws):
    """读取工作表全部数据，处理好合并单元格"""
    # 先记下合并信息
    merged_map = {}
    for mr in ws.merged_cells.ranges:
        val = _get_cell_value(ws.cell(mr.min_row, mr.min_col))
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_map[(r, c)] = val

    data = []
    for row_idx in range(1, ws.max_row + 1):
        row_data = []
        for col_idx in range(1, max(ws.max_column, 1) + 1):
            if (row_idx, col_idx) in merged_map:
                row_data.append(merged_map[(row_idx, col_idx)])
            else:
                val = _get_cell_value(ws.cell(row_idx, col_idx))
                row_data.append(val)
        data.append(row_data)
    return data


def _to_text(data):
    """将二维数据转换为格式化的文本"""
    lines = []
    for row in data:
        # 过滤空行
        cells = [str(c).strip().replace("\u201c","").replace("\u201d","") if c else "" for c in row]
        if any(c for c in cells):
            lines.append(" | ".join(cells))
    return "\n".join(lines)


def analyze_table(table_id):
    """分析表格，返回 [{'table_id': ..., 'sheet_name': ..., 'text': ..., 'data': [[...]]}]"""
    path = get_table_path(table_id)
    wb = openpyxl.load_workbook(path, data_only=True)
    results = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        raw = _read_sheet_data(ws)
        text = _to_text(raw)
        if text.strip():
            results.append({
                "table_id": table_id,
                "sheet_name": sn,
                "text": text,
                "data": raw,
            })
    wb.close()
    return results
